import openpyxl
from openpyxl.styles import PatternFill # to fill color in excel import Pattern Fill class

def getRowCount(file, SheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[SheetName]
    return(sheet.max_row)

def getColumnCount(file, SheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[SheetName]
    return(sheet.max_column)

def readData(file, SheetName, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[SheetName]
    return sheet.cell(rownum, columnno).value

def writeData(file, SheetName, rownum, columnno, data):
    workbook= openpyxl.load_workbook(file)
    sheet = workbook[SheetName]
    sheet.cell(rownum, columnno).value = data
    workbook.save(file)

def fillGreenColor(file, SheetName, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[SheetName]
    greenFill = PatternFill(start_color="60b212",
                            end_color="60b212",
                            fill_type="solid")
    sheet.cell(rownum, columnno).fill=greenFill
    workbook.save(file)


def fillRedColor(file, SheetName, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[SheetName]
    redFill = PatternFill(start_color="ff0000",
                            end_color="ff0000",
                            fill_type="solid")
    sheet.cell(rownum, columnno).fill=redFill
    workbook.save(file)